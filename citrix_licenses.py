import random
import openpyxl
import datetime
import pymupdf


POINTS_PER_MM = 2.83464567
PAGEWIDTH = 840 * POINTS_PER_MM
PAGEHEIGHT = 594 * POINTS_PER_MM

TB_MARGIN = 15 * POINTS_PER_MM
T_HEADER = 5 * POINTS_PER_MM
LR_MARGIN = 15 * POINTS_PER_MM
L_HEADER = 20 * POINTS_PER_MM

SESSION_HEIGHT = 2 * POINTS_PER_MM
SESSION_MARGIN = 0.5 * POINTS_PER_MM
NSESSIONS_PER_SERVER = 5
SESSION_COLORS = {"MonacoDG": (1, 0, 0), "MonacoMRL": (0, 0, 1), "MonacoSim": (0, 1, 0), "MonacoDev": (1, 0, 1), "Total": (0,)}

CHART_WIDTH = PAGEWIDTH - 2 * LR_MARGIN - L_HEADER
CHART_HEIGHT = 45 * POINTS_PER_MM
CHART_MARGIN = 6 * POINTS_PER_MM
USERNAME_INTERVAL = 15 * POINTS_PER_MM

FONT_ARCHIVE = pymupdf.Archive("C:/Windows/Fonts")
headers = ["User", "Server", "SessionStartTime", "SessionEndTime", "SessionDuration"]


def hours_since(time, reference):
    dt = time - reference
    return dt/datetime.timedelta(hours=1)


class CitrixSession(object):
    def __init__(self, user, server, environment, start_time, end_time):
        self.user = user
        self.server = server
        self.environment = environment
        self.start_time = start_time
        self.end_time = end_time

    def __lt__(self, other):
        return self.start_time < other.start_time

    def __gt__(self, other):
        return self.start_time > other.start_time

    def __eq__(self, other):
        return self.start_time == other.start_time

    @property
    def hash(self):
        return str(hash(f"{self.user}{self.server}{self.environment}{self.start_time}"))

    @property
    def duration(self):
        if self.end_time is None:
            return None
        return (self.end_time - self.start_time)/datetime.timedelta(minutes=1)

    def __str__(self):
        return f"CitrixSession({self.user}@{self.server} {self.start_time} - {self.end_time})"

    def __repr__(self):
        return str(self)


class CitrixSessionStart(object):
    def __init__(self, session):
        self.timestamp = session.start_time
        self.session = session
        self.start = True
        self.end = False


class CitrixSessionEnd(object):
    def __init__(self, session):
        self.timestamp = session.end_time
        self.session = session
        self.start = False
        self.end = True


def session_events(sessions):
    """Returns a list of events starting or stopping a session."""
    session_events = []
    end_time = max([s.end_time for s in sessions.values() if s.end_time is not None]) + datetime.timedelta(minutes=1)
    for s in sessions.values():
        if s.end_time is None:
            s.end_time = end_time
        session_events.append(CitrixSessionStart(s))
        session_events.append(CitrixSessionEnd(s))
    sorted_session_events = sorted(session_events, key=lambda s: s.timestamp)
    return sorted_session_events


def sessions_per_user(sessions):
    users = set([s.user for s in sessions.values()])
    sorted_session_events = session_events(sessions)
    start_time = sorted_session_events[0].timestamp
    xs = [start_time]
    ysdict = {u: [0] for u in users}

    offenders = set()

    for s in sorted_session_events:
        xs.append(s.timestamp)
        for user in ysdict.keys():
            val = ysdict[user][-1]
            if user == s.session.user:
                if s.start:
                    val += 1
                else:
                    val -= 1
                if val > 3:
                    offenders.add(user)
            ysdict[user].append(val)

    return xs, ysdict, offenders


def average_sessions_per_user(xs, ysdict):
    ys = []

    for i in range(len(xs)):
        val = 0
        n_active_users = 0
        for user in ysdict.keys():
            user_sessions = ysdict[user][i]
            if user_sessions > 0:
                n_active_users += 1
            val += ysdict[user][i]

        if n_active_users > 0:
            val /= n_active_users

        ys.append(val)
    return xs, ys


def concurrent_sessions(sessions):
    sorted_session_events = session_events(sessions)
    start_time = sorted_session_events[0].timestamp
    xs = [start_time]
    ys = {env: [0] for env in SESSION_COLORS.keys()}

    sessions_used = {env: 0 for env in SESSION_COLORS.keys()}
    for session_event in sorted_session_events:
        env = session_event.session.environment
        if session_event.start:
            sessions_used[env] += 1
            sessions_used["Total"] += 1
        else:
            sessions_used[env] -= 1
            sessions_used["Total"] -= 1
        xs.append(session_event.timestamp)
        for env in SESSION_COLORS.keys():
            ys[env].append(sessions_used[env])

    return xs, ys


def concurrent_users(sessions):
    sorted_session_events = session_events(sessions)
    start_time = sorted_session_events[0].timestamp
    usernames = set([s.user for s in sessions.values()])
    xs = [start_time]
    ys = [0]
    active_users = {u: 0 for u in usernames}

    for session_event in sorted_session_events:
        if session_event.start:
            active_users[session_event.session.user] += 1
        else:
            active_users[session_event.session.user] -= 1

        xs.append(session_event.timestamp)
        ys.append(sum([1 for u in active_users.items() if u[1] > 0]))

    return xs, ys


def draw_rect(page, x1, y1, x2, y2, width=0.1, color=(1,0,0), fill=None):
    rect = pymupdf.Rect(x1, y1, x2, y2)
    page.draw_rect(rect, fill=fill, color=color, width=width)


def draw_line(page, x1, y1, x2, y2, width=0.1, color=(0,)):
    p1 = pymupdf.Point(x1, y1)
    p2 = pymupdf.Point(x2, y2)
    page.draw_line(p1, p2, width=width, color=color)


def add_text(page, x, y, text, fontsize=9, color=(0,), rotate=0):
    p = pymupdf.Point(x, y)
    page.insert_text(p, text, color=color, fontsize=fontsize, fontname="GilSansMT", fontfile="C:/Windows/Fonts/GIL_____.ttf", rotate=rotate)


def sample_and_hold_graph(xs, ys):
    xx = [xs[0]]
    yy = [ys[0]]

    for x, y in list(zip(xs, ys))[1:]:
        xx.append(x)
        xx.append(x)
        yy.append(yy[-1])
        yy.append(y)

    return xx, yy


def simple_graph(page, top, title, xs, ys, xmin, xmax, tickinterval=5, color=(0,), width=0.1):
    ymax = max(ys)
    deltax = draw_graph(page, top, title, xmin, xmax, ymax, tickinterval)

    xxs = [LR_MARGIN + L_HEADER + CHART_WIDTH * hours_since(x, xmin) / deltax for x in xs]
    yys = [top + CHART_HEIGHT * (1 - y / ymax) for y in ys]
    points = [pymupdf.Point(x, y) for x, y in zip(xxs, yys)]
    page.draw_polyline(points, color=color, width=width)


def multiple_graph(page, top, title, xs, ysdict, xmin, xmax, legend_labels=None, colors= None, tickinterval=5, width=0.1):
    ymax = max([max(v) for v in ysdict.values()])
    deltax = draw_graph(page, top, title, xmin, xmax, ymax, tickinterval)

    n = 0
    row = 0
    for label, ys in ysdict.items():
        if colors is None:
            color = tuple(random.random() for _ in range(3))
        else:
            color = colors[label]
        xxs = [LR_MARGIN + L_HEADER + CHART_WIDTH * hours_since(x, xmin) / deltax for x in xs]
        yys = [top + CHART_HEIGHT * (1 - y / ymax) for y in ys]
        points = [pymupdf.Point(x, y) for x, y in zip(xxs, yys)]
        page.draw_polyline(points, color=color, width=width)
        if legend_labels is not None and label in legend_labels:
            add_text(page, LR_MARGIN + L_HEADER + 2 + n*USERNAME_INTERVAL, top + 5 + row * 7, label, fontsize=5, color=color)
            n += 1
            if n > CHART_WIDTH/USERNAME_INTERVAL:
                n = 0
                row += 1


def draw_graph(page, top, title, xmin, xmax, ymax, tickinterval=5):
    add_text(page, LR_MARGIN + 0.75*L_HEADER, top + CHART_HEIGHT, title, rotate=90, fontsize=6)
    draw_line(page, LR_MARGIN + L_HEADER, top, LR_MARGIN + L_HEADER, top + CHART_HEIGHT)
    draw_line(page, LR_MARGIN + L_HEADER, top + CHART_HEIGHT, LR_MARGIN + L_HEADER + CHART_WIDTH, top + CHART_HEIGHT)

    deltax = (xmax - xmin) / datetime.timedelta(hours=1)

    tick = 0
    while tick <= ymax:
        tickloc = top + CHART_HEIGHT * (1 - tick / ymax)
        add_text(page, LR_MARGIN + 47, tickloc + 2, f"{tick}", fontsize=5)
        draw_line(page, LR_MARGIN + L_HEADER - 3, tickloc, LR_MARGIN + L_HEADER, tickloc)
        draw_line(page, LR_MARGIN + L_HEADER, tickloc, LR_MARGIN + L_HEADER + CHART_WIDTH, tickloc, color=(0.8,))
        tick += tickinterval

    current_day = xmin
    while current_day <= xmax:
        t = current_day.strftime("%a %d/%m")
        l = CHART_WIDTH * hours_since(current_day, xmin) / deltax + LR_MARGIN + L_HEADER
        draw_line(page, l, top, l, top + CHART_HEIGHT, color=(0.8,))
        current_day += datetime.timedelta(days=1)
        if current_day <= xmax:
            daywidth = CHART_WIDTH * 24 / deltax
            add_text(page, l + 0.1 * daywidth, top + CHART_HEIGHT + 7, t, fontsize=5)

    return deltax


def session_plot(sessions):
    """"""
    doc = pymupdf.open()  # create empty PDF
    page = doc.new_page(width=PAGEWIDTH, height=PAGEHEIGHT)  # create an empty A4 page

    print("Sorting sessions")
    sorted_session_events = session_events(sessions)
    start_time = sorted_session_events[0].timestamp
    start_time = start_time.replace(hour=0, minute=0, second=0, microsecond=0)
    end_time = max([s.timestamp for s in sorted_session_events])
    end_time = end_time.replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=1)
    total_time = (end_time - start_time)/datetime.timedelta(hours=1)

    servers = sorted(set([s.server for s in sessions.values()]))

    if True:
        print("Creating session plot")
        server_sessions = {s:[None for i in range(NSESSIONS_PER_SERVER)] for s in servers}

        current_day = start_time
        while current_day <= end_time:
            t = current_day.strftime("%a %d/%m")
            l = CHART_WIDTH * hours_since(current_day, start_time) / total_time + LR_MARGIN + L_HEADER
            draw_line(page, l, TB_MARGIN, l, TB_MARGIN + T_HEADER + len(servers) * NSESSIONS_PER_SERVER * (SESSION_HEIGHT + SESSION_MARGIN), color=(0.5,))
            current_day += datetime.timedelta(days=1)
            if current_day <= end_time:
                daywidth = CHART_WIDTH * 24/total_time
                add_text(page, l+0.1*daywidth, TB_MARGIN + 0.5*T_HEADER, t, fontsize=5)

        for n, server in enumerate(servers):
            row = n * NSESSIONS_PER_SERVER
            y = TB_MARGIN + T_HEADER + row *(SESSION_HEIGHT + SESSION_MARGIN) + SESSION_HEIGHT
            add_text(page, LR_MARGIN, y, server[6:], fontsize=9)

        print("\tSorting data")
        sorted_session_events = session_events(sessions)
        double_sorted_session_events = sorted(sorted_session_events, key= lambda x: (x.session.server, x.timestamp))

        row_offsets = {}

        print("\tCreating diagram")
        for s in double_sorted_session_events:
            if s.start:
                server_idx = servers.index(s.session.server)
                server_session_idx = server_sessions[s.session.server].index(None)
                server_sessions[s.session.server][server_session_idx] = s.session.hash
                row = server_idx * NSESSIONS_PER_SERVER + server_session_idx
                if row not in row_offsets:
                    row_offsets[row] = -2

                l = CHART_WIDTH * hours_since(s.timestamp, start_time) / total_time + LR_MARGIN + L_HEADER
                r = CHART_WIDTH * hours_since(s.session.end_time, start_time) / total_time + LR_MARGIN + L_HEADER
                t = TB_MARGIN + T_HEADER + row * (SESSION_HEIGHT + SESSION_MARGIN)
                b = t + SESSION_HEIGHT
                draw_rect(page, l, t, r, b, color=SESSION_COLORS[s.session.environment])
                vpos = 0.5*(t+b) + 0.5 + row_offsets[row]
                row_offsets[row] += 1.33333
                if row_offsets[row] > 2:
                    row_offsets[row] = -2
                add_text(page, l+1, vpos, s.session.user, fontsize=2, rotate=0)
            else:
                server_sessions[s.session.server][server_sessions[s.session.server].index(s.session.hash)] = None


    # Graph of number of sessions in use
    print("Create number of sessions graph")
    graph_top = len(servers) * NSESSIONS_PER_SERVER * (SESSION_HEIGHT + SESSION_MARGIN) + TB_MARGIN + T_HEADER + CHART_MARGIN
    print("\tCompiling data")
    csx, csydict = concurrent_sessions(sessions)

    print("\tConvert to step graph")
    csyydict = {}
    for env, csy in csydict.items():
        csxx, csyy = sample_and_hold_graph(csx, csy)
        csyydict[env] = csyy
    print("\tCreating diagram")
    multiple_graph(page, graph_top, "Number of sessions in use", csxx, csyydict, start_time, end_time, colors=SESSION_COLORS)

    # Graph of users
    print("Create user graph")
    graph_top = len(servers) * NSESSIONS_PER_SERVER * (SESSION_HEIGHT + SESSION_MARGIN) + TB_MARGIN + T_HEADER + CHART_HEIGHT + 2 * CHART_MARGIN
    print("\tCompiling data")
    ux, uy = concurrent_users(sessions)
    print("\tConvert to step graph")
    uxx, uyy = sample_and_hold_graph(ux, uy)
    print("\tCreating diagram")
    simple_graph(page, graph_top, "Number of concurrent users", uxx, uyy, start_time, end_time)

    # Graph of sessions per user
    print("Create sessions per user graph")
    graph_top = len(servers) * NSESSIONS_PER_SERVER * (SESSION_HEIGHT + SESSION_MARGIN) + TB_MARGIN + T_HEADER + 2 * CHART_HEIGHT + 3 * CHART_MARGIN
    print("\tCompiling data")
    usx, usydict, offenders = sessions_per_user(sessions)
    print("\tConvert to step graph")
    usyydict = {}
    for user, usy in usydict.items():
        usxx, usyy = sample_and_hold_graph(usx, usy)
        usyydict[user] = usyy
    print("\tCreating diagram")
    multiple_graph(page, graph_top, "Number of sessions per user", usxx, usyydict, start_time, end_time, legend_labels=offenders, tickinterval=1)

    print("Create average sessions per user graph")
    graph_top = len(servers) * NSESSIONS_PER_SERVER * (SESSION_HEIGHT + SESSION_MARGIN) + TB_MARGIN + T_HEADER + 3 * CHART_HEIGHT + 4 * CHART_MARGIN
    print("\tCompiling data")
    usx, usy = average_sessions_per_user(usx, usydict)
    print("\tConvert to step graph")
    usxx, usyy = sample_and_hold_graph(usx, usy)
    print("\tCreating diagram")
    simple_graph(page, graph_top, "Average sessions per active user", usxx, usyy, start_time, end_time, tickinterval=0.5)

    doc.save("output.pdf")


def main():
    sessions = {}
    user_sessions = {}

    print("Loading data")
    wb = openpyxl.load_workbook('U:\Monaco\Licenties\AlleSessies.xlsx')
    sheet = wb.active

    for row in range(2, sheet.max_row):
        user = sheet.cell(row=row, column=1).value
        server = sheet.cell(row=row, column=2).value
        environment = sheet.cell(row=row, column=3).value

        start_time = datetime.datetime.strptime(sheet.cell(row=row, column=4).value, "%m/%d/%Y %I:%M:%S %p")

        try:
            end_time = datetime.datetime.strptime(sheet.cell(row=row, column=5).value, "%m/%d/%Y %I:%M:%S %p")
        except ValueError:
            end_time = None

        s = CitrixSession(user, server, environment, start_time, end_time)

        if s.hash in sessions:
            if sessions[s.hash].end_time is None:
                sessions[s.hash] = s
            elif s.end_time is None:
                pass
            elif sessions[s.hash].end_time != s.end_time:
                raise ValueError(f"Session already exists: {s}")
        else:
            sessions[s.hash] = s

        if user not in user_sessions.keys():
            user_sessions[user] = []
        user_sessions[user].append(s)


    session_plot(sessions)



if __name__ == "__main__":
    main()